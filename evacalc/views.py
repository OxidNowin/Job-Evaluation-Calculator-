import os

from django.http import Http404
from django.http import FileResponse
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required

from .models import JobEvaluation
from .forms import PostForm, UnlogicalPostForm
from .arrays import skills_arr, responsibility_arr, problems_solving_arr, union_arr, grade_arr

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, GradientFill


@login_required
def result(request):
    return render(request, 'evacalc/result.html')


@login_required
def archive_date(request):
    if request.method == "POST":
        job_dict = request.POST
        job_title = list(job_dict.values())[1]
        redirect_str = "archive_date"
        delete_job_evaluation(job_title, redirect_str)
    jobs_list = JobEvaluation.objects.order_by('-created_date').filter(user=request.user)
    context = {'jobs_list': jobs_list}
    return render(request, 'evacalc/archive.html', context)


@login_required
def archive_grade(request):
    if request.method == "POST":
        job_dict = request.POST
        job_title = list(job_dict.values())[1]
        redirect_str = "archive_grade"
        delete_job_evaluation(job_title, redirect_str)
    jobs_list = JobEvaluation.objects.order_by('-grade').filter(user=request.user)
    context = {'jobs_list': jobs_list}
    return render(request, 'evacalc/archive.html', context)


@login_required
def index(request):
    if request.method == "POST" and 'logical_post' in request.POST:
        form = PostForm(request.POST)
        if form.is_valid():

            """Считывание данных"""

            title = form.cleaned_data['title'].title()
            short_profile = form.cleaned_data['short_profile']
            hard_skills = str(form.cleaned_data['hard_skills'].upper())
            knowledge = str(form.cleaned_data['knowledge'].upper())
            soft_skills = str(form.cleaned_data['soft_skills'].upper())
            around_question = str(form.cleaned_data['around_question'].upper())
            question_complexity = str(form.cleaned_data['question_complexity'].upper())
            freedom_action = str(form.cleaned_data['freedom_action'].upper())
            nature_impact = str(form.cleaned_data['nature_impact'].upper())
            impact_importance = str(form.cleaned_data['impact_importance'].upper())

            """Вычисления"""

            value_of_skills_section, is_skills_section_logical = compute_skills_section(hard_skills, knowledge,
                                                                                        soft_skills)
            value_of_problems_section, is_problems_section_logical = compute_problems_section(around_question,
                                                                                              question_complexity)
            value_of_union_section, is_union_of_sections_logical = compute_union_skills_and_problems(
                value_of_skills_section, value_of_problems_section)
            value_of_responsibility_section = compute_responsibility_section(freedom_action, nature_impact,
                                                                             impact_importance)

            """Проверка на ненахождение соответствия"""

            if is_none(value_of_skills_section, value_of_problems_section, value_of_union_section,
                       value_of_responsibility_section):
                return render(request, "evacalc/index.html", {'form': form,
                                                              'error_message': 'Неккоректно ввели значения', })

            if is_waterfall_principle(hard_skills, around_question, freedom_action) is False:
                return render(request, "evacalc/index.html", {'form': form,
                                                              'error_message': 'Не соблюден принцип водопада', })

            """ Суммирование значений """

            sum_of_values = value_of_skills_section + value_of_union_section + value_of_responsibility_section
            grade = grade_determine(sum_of_values)

            result_str = f"{request.user} {title} {short_profile} {hard_skills} {knowledge} {soft_skills} {value_of_skills_section} {around_question} {question_complexity} {int(value_of_problems_section * 100)} {value_of_union_section} {freedom_action} {nature_impact} {impact_importance} {value_of_responsibility_section} {sum_of_values} {grade} "

            """Проверка на логичность"""

            if (is_skills_section_logical and is_problems_section_logical and is_union_of_sections_logical) is False:
                return render(request, "evacalc/index.html", {'form': form,
                                                              'logical_message': 'Нелогичность данных. Продолжить?',
                                                              'result_str': result_str})
            job_evaluation_save = JobEvaluation(
                title=title,
                user=request.user,
                short_profile=short_profile,
                hard_skills=hard_skills,
                knowledge=knowledge,
                soft_skills=soft_skills,
                value_of_skills_section=value_of_skills_section,
                around_question=around_question,
                question_complexity=question_complexity,
                value_of_problems_section=int(value_of_problems_section * 100),
                value_of_union_section=value_of_union_section,
                freedom_action=freedom_action,
                nature_impact=nature_impact,
                impact_importance=impact_importance,
                value_of_responsibility_section=value_of_responsibility_section,
                sum_of_values=value_of_skills_section + value_of_union_section + value_of_responsibility_section,
                grade=grade)
            job_evaluation_save.save()
            form = PostForm()
            return render(request, 'evacalc/index.html', {'form': form,
                                                          'success': result_str})
        else:
            return render(request, 'evacalc/index.html', {'form': form,
                                                          'error_message': 'Не указали наименование должности'})
    elif request.method == "POST" and 'unlogical_post' in request.POST:
        form = UnlogicalPostForm(request.POST)
        if form.is_valid():
            result_str = form.cleaned_data['unlogical_result']
            result_arr = result_str.strip().split()
            print(result_arr)
            job_evaluation_save = JobEvaluation(
                title=result_arr[1],
                user=request.user,
                short_profile=result_arr[2],
                hard_skills=result_arr[3],
                knowledge=result_arr[4],
                soft_skills=result_arr[5],
                value_of_skills_section=result_arr[6],
                around_question=result_arr[7],
                question_complexity=result_arr[8],
                value_of_problems_section=int(result_arr[9]),
                value_of_union_section=result_arr[10],
                freedom_action=result_arr[11],
                nature_impact=result_arr[12],
                impact_importance=result_arr[13],
                value_of_responsibility_section=result_arr[14],
                sum_of_values=result_arr[15],
                grade=result_arr[16])
            job_evaluation_save.save()
            form = PostForm()
            return render(request, 'evacalc/index.html', {"form": form,
                                                          'success': result_str})
    else:
        form = PostForm()
    return render(request, "evacalc/index.html", {'form': form})


def compute_skills_section(hard_skills, knowledge, soft_skills):

    """Практические занятия Управленческие знания Навыки взаимодействия"""

    is_skills_section_logical = True

    for arr in skills_arr:
        if str(arr[0]) == hard_skills:
            if str(arr[1]) == knowledge:
                if str(arr[2]) == soft_skills:
                    if str(arr[4]) == "0":
                        is_skills_section_logical = False
                    return int(arr[3]), is_skills_section_logical
    return None, None


def compute_problems_section(around_question, question_complexity):

    """Область решаемых вопросов Сложность вопросов"""

    is_problems_section_logical = True

    for arr in problems_solving_arr:
        if str(arr[0]) == around_question:
            if str(arr[1]) == question_complexity:
                if str(arr[3]) == "0":
                    is_problems_section_logical = False
                return float(arr[2]), is_problems_section_logical
    return None, None


def compute_union_skills_and_problems(value_of_skills_section, value_of_problems_section):
    is_union_of_sections_logical = True

    for arr in union_arr:
        if str(arr[0]) == str(value_of_problems_section):
            if str(arr[1]) == str(value_of_skills_section):
                if str(arr[3]) == "0":
                    is_union_of_sections_logical = False
                return int(arr[2]), is_union_of_sections_logical
    return None, None


def compute_responsibility_section(freedom_action, nature_impact, impact_importance):

    """ Свобода действий Природа воздействия Важность воздействия"""

    for arr in responsibility_arr:
        if str(arr[0]) == freedom_action:
            if str(arr[1]) == nature_impact:
                if str(arr[2]) == impact_importance:
                    return int(arr[3])
    return None


def grade_determine(sum_of_values):
    for summ in grade_arr:
        if (sum_of_values <= summ[2]) and (sum_of_values >= summ[1]):
            return summ[0]
    return None


def is_none(value_of_skills_section, value_of_problems_section, value_of_union_section,
            value_of_responsibility_section):
    if (value_of_skills_section is None) or (value_of_problems_section is None) or (value_of_union_section is None) or (
            value_of_responsibility_section is None):
        return True
    return False


def is_waterfall_principle(hard_skills, around_question, freedom_action):
    if ord(hard_skills) < ord(around_question):
        return False
    elif ord(around_question) < ord(freedom_action):
        return False
    return True


def delete_job_evaluation(title, redirect_str):
    job_name = JobEvaluation.objects.get(title=title)
    job_name.delete()
    return redirect(redirect_str)


def create_xl(request):
    jobs_dicts = list(JobEvaluation.objects.filter(user=request.user).values())
    filename = f'{request.user}JobEvaluation.xlsx'
    jobs_arr = []
    for dictionary in jobs_dicts:
        jobs_arr.append([])
        jobs_arr[-1].append(dictionary['title'])
        jobs_arr[-1].append(dictionary['short_profile'])
        jobs_arr[-1].append(dictionary['hard_skills'])
        jobs_arr[-1].append(dictionary['knowledge'])
        jobs_arr[-1].append(dictionary['soft_skills'])
        jobs_arr[-1].append(dictionary['value_of_skills_section'])
        jobs_arr[-1].append(dictionary['around_question'])
        jobs_arr[-1].append(dictionary['question_complexity'])
        jobs_arr[-1].append(dictionary['value_of_problems_section'])
        jobs_arr[-1].append(dictionary['value_of_union_section'])
        jobs_arr[-1].append(dictionary['freedom_action'])
        jobs_arr[-1].append(dictionary['nature_impact'])
        jobs_arr[-1].append(dictionary['impact_importance'])
        jobs_arr[-1].append(dictionary['value_of_responsibility_section'])
        jobs_arr[-1].append(dictionary['sum_of_values'])
        jobs_arr[-1].append(dictionary['grade'])
    wb = Workbook()
    ws = wb.active
    ws.title = "Должности"

    """ Заголовок """

    ws.merge_cells('B1:Z2')
    ws['B1'] = "Калькулятор оценки должностей EPSI Rating"
    thin = Side(border_style="thin", color="000000")
    ws['B1'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
    ws['B1'].fill = PatternFill("solid", fgColor="DDDDDD")
    ws['B1'].font = Font(name='TimesNewRoman', size=22, b=True, color="FF0000")
    ws['B1'].alignment = Alignment(horizontal="center", vertical="center")

    """ Заголовки ячеек """

    letter_numb = 65
    for _ in range(13):
        ws.merge_cells(f'{chr(letter_numb)}3:{chr(letter_numb + 1)}4')
        letter_numb += 2
    ws.merge_cells('AA3:AA4')
    ws['A3'] = "Название должности"
    ws['C3'] = "Краткий профиль"
    ws['E3'] = "Практические знания"
    ws['G3'] = "Управленческие знания"
    ws['I3'] = "Навыки общения"
    ws['K3'] = "Пункт оценки"
    ws['M3'] = "Область вопросов"
    ws['O3'] = "Сложность вопросов"
    ws['Q3'] = "Значение в %"
    ws['S3'] = "Свобода действий"
    ws['U3'] = "Природа воздействия"
    ws['W3'] = "Важность воздействия"
    ws['Y3'] = "Пункты оценки"
    ws['AA3'] = "Грейд"
    letter_numb = 65
    for i in range(13):
        xl_title = ws[f'{chr(letter_numb)}3']
        if i % 2 == 0:
            bg_color = "E41717"
        else:
            bg_color = "B32D2D"
        xl_title.border = Border(top=thin, left=thin, right=thin, bottom=thin)
        xl_title.fill = PatternFill("solid", fgColor=bg_color)
        xl_title.font = Font(name='TimesNewRoman', size=14, b=True, color="000000")
        xl_title.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
        letter_numb += 2
    last_cell = ws['AA3']
    last_cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    last_cell.fill = PatternFill("solid", fgColor="B32D2D")
    last_cell.font = Font(name='TimesNewRoman', size=14, b=True, color="000000")
    last_cell.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")

    """ Значения """

    row_number = 5
    for each in jobs_arr:
        letter_numb = 65
        for i in range(13):
            cell = ws[f'{chr(letter_numb)}{row_number}']
            ws.merge_cells(f'{chr(letter_numb)}{row_number}:{chr(letter_numb + 1)}{row_number}')
            ws[f'{chr(letter_numb)}{row_number}'] = each[i]
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            cell.font = Font(name='TimesNewRoman', size=12, b=True, color="000000")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            letter_numb += 2
        last_cell = ws[f"AA{row_number}"]
        ws[f"AA{row_number}"] = each[-1]
        last_cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
        last_cell.font = Font(name='TimesNewRoman', size=12, b=True, color="000000")
        last_cell.alignment = Alignment(horizontal="center", vertical="center")
        row_number += 1
    wb.save(filename=filename)


def get_file(request):
    create_xl(request)
    file_path = f'{request.user}JobEvaluation.xlsx'
    if os.path.exists(file_path):
        response = FileResponse(open(file_path, 'rb'))
        return response
    raise Http404
